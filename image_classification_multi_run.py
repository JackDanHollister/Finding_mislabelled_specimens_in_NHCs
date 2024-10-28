# %%
# Read the iteration number from the file
with open("iteration.txt", "r") as f:
    iteration = int(f.read())


# %%
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'  # or set it to '3' to suppress all messages, including INFO and WARNING



# %%
# Import libraries
import os
import random
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import shutil
from sklearn.metrics import confusion_matrix, classification_report
import keras_tuner as kt

import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import layers, backend, optimizers, regularizers
from tensorflow.keras.preprocessing.image import ImageDataGenerator, load_img, img_to_array
from tensorflow.keras.models import Sequential, Model
from tensorflow.keras.callbacks import ModelCheckpoint, LearningRateScheduler, TensorBoard, EarlyStopping
from tensorflow.keras.applications.vgg16 import VGG16, preprocess_input
from tensorflow.keras.layers import Conv2D, MaxPooling2D, ZeroPadding2D, Activation, Flatten, Dense, Dropout, GlobalAveragePooling2D, BatchNormalization


# %%
from matplotlib import cm
from tf_keras_vis.gradcam import Gradcam

from tf_keras_vis.utils.model_modifiers import ReplaceToLinear
replace2linear = ReplaceToLinear()

# %%


# %%
import tensorflow as tf
print("Number of GPUs Available: ", len(tf.config.experimental.list_physical_devices('GPU')))


# %%
SEED = 666

#Function to initialize seeds for all libraries which might have stochastic behavior
def set_seeds(seed=SEED):
    os.environ['PYTHONHASHSEED'] = str(seed)
    random.seed(seed)
    tf.random.set_seed(seed)
    np.random.seed(seed)

# Activate Tensorflow deterministic behavior
def set_global_determinism(seed=SEED):
    set_seeds(seed=seed)

    os.environ['TF_DETERMINISTIC_OPS'] = '1'
    os.environ['TF_CUDNN_DETERMINISTIC'] = '1'
    
    tf.config.threading.set_inter_op_parallelism_threads(1)
    tf.config.threading.set_intra_op_parallelism_threads(1)

# Call the above function with seed value
set_global_determinism(seed=SEED)

# %%
from datetime import datetime

date = datetime.now().strftime('%Y_%m_%d-%I:%M_%S_%p')

# %%


# Define the path to the root directory where the subdirectories are located
root_dir = "1butt_bbox_cut_minus_1_orig/"


# Define the path to the new directory to create
new_dir = "save_offs/pt1"



# %%


# %%


# %% [markdown]
# # COPY OVER SUBDIRECTORIES IF THEY ARE >=200

# %%
# Create the new directory if it does not exist
if not os.path.exists(new_dir):
    os.makedirs(new_dir)

# Iterate through the subdirectories in the root directory
for subdir in os.listdir(root_dir):
    # Construct the full path to the subdirectory
    subdir_path = os.path.join(root_dir, subdir)
    
    # Check if the subdirectory contains at least 200 images
    if len(os.listdir(subdir_path)) >= 200:
        # Copy the subdirectory and its contents to the new directory
        shutil.copytree(subdir_path, os.path.join(new_dir, subdir))


# %% [markdown]
# # orangise these new images into train, val and test

# %%
import os
import random
import shutil

# Define source directory and destination directory paths
src_dir = "save_offs/pt1"  # Update to the correct source directory path
dst_dir = "save_offs/pt2"

# Create train, validation, and test directories if they do not exist
train_dir = os.path.join(dst_dir, "train")
val_dir = os.path.join(dst_dir, "validation")
test_dir = os.path.join(dst_dir, "test")
for d in [train_dir, val_dir, test_dir]:
    if not os.path.exists(d):
        os.makedirs(d)

# Loop through subdirectories in source directory
for subdir in os.listdir(src_dir):
    subdir_path = os.path.join(src_dir, subdir)
    if os.path.isdir(subdir_path):
        # Create subdirectories in train, validation, and test directories with identical names
        train_subdir = os.path.join(train_dir, subdir)
        val_subdir = os.path.join(val_dir, subdir)
        test_subdir = os.path.join(test_dir, subdir)
        for d in [train_subdir, val_subdir, test_subdir]:
            if not os.path.exists(d):
                os.makedirs(d)

        # Get list of image files in the subdirectory
        image_files = [f for f in os.listdir(subdir_path) if f.endswith(".JPG")]

        # Specify number of training and validation samples
        num_train_samples = min(250, len(image_files))
        num_val_samples = min(50, len(image_files) - num_train_samples)

        # Randomly sample images for train, validation, and test sets
        num_total_samples = len(image_files)
        num_test_samples = num_total_samples - num_train_samples - num_val_samples
        train_samples = random.sample(image_files, num_train_samples)
        remaining_files = list(set(image_files) - set(train_samples))
        val_samples = random.sample(remaining_files, num_val_samples)
        remaining_files = list(set(remaining_files) - set(val_samples))
        test_samples = random.sample(remaining_files, num_test_samples)

        # Move images to train, validation, and test subdirectories
        for train_sample in train_samples:
            src_path = os.path.join(subdir_path, train_sample)
            dst_path = os.path.join(train_subdir, train_sample)
            shutil.move(src_path, dst_path)

        for val_sample in val_samples:
            src_path = os.path.join(subdir_path, val_sample)
            dst_path = os.path.join(val_subdir, val_sample)
            shutil.move(src_path, dst_path)

        for test_sample in test_samples:
            src_path = os.path.join(subdir_path, test_sample)
            dst_path = os.path.join(test_subdir, test_sample)
            shutil.move(src_path, dst_path)


# %%


# %%


# %% [markdown]
# # create a seperate test_even amount

# %%


# Create a new directory called "test_even" within the destination directory
test_dir = os.path.join(dst_dir, "test")
test_even_dir = os.path.join(dst_dir, "test_even")
if not os.path.exists(test_even_dir):
    os.makedirs(test_even_dir)

# Loop through subdirectories in test directory
for subdir in os.listdir(test_dir):
    subdir_path = os.path.join(test_dir, subdir)
    if os.path.isdir(subdir_path):
        # Create identical subdirectories within test_even directory
        test_even_subdir = os.path.join(test_even_dir, subdir)
        if not os.path.exists(test_even_subdir):
            os.makedirs(test_even_subdir)

        # Randomly sample 20 images from the subdirectory
        image_files = [f for f in os.listdir(subdir_path) if f.endswith(".JPG")]
        num_images = len(image_files)
        if num_images < 100:
            num_samples = num_images
        else:
            num_samples = 100
        sample_files = random.sample(image_files, num_samples)

        # Copy sampled images to the test_even subdirectory
        for sample_file in sample_files:
            src_path = os.path.join(subdir_path, sample_file)
            dst_path = os.path.join(test_even_subdir, sample_file)
            shutil.copy(src_path, dst_path)


# %%


# %%
import os
import Augmentor

dir = 'save_offs/pt2/train/'

# Get a list of all subdirectories within the main directory
subdirs = [d for d in os.listdir(dir) if os.path.isdir(os.path.join(dir, d))]

# Loop over each subdirectory and apply augmentations
for subdir in subdirs:
    subdir_path = os.path.join(dir, subdir)
    p = Augmentor.Pipeline(subdir_path, output_directory='')

    p.flip_left_right(probability=0.8)
    p.flip_top_bottom(probability=0.8)
    #p.rotate(probability=0.9,max_left_rotation=18,max_right_rotation=18)
    p.rotate90(probability=0.7)
    p.rotate180(probability=0.7)
    p.rotate270(probability=0.7)
    #p.scale(probability=0.7, scale_factor=1.1)
    p.random_brightness(probability=0.7,min_factor=0.8,max_factor=1.2)
    p.random_contrast(probability=0.7,min_factor=0.8,max_factor=1.2)

    # Get the number of files in the directory
    num_files = len(os.listdir(subdir_path))

    # Sample additional images if necessary
    num_samples = 1000 - num_files
    if num_samples > 0:
        p.sample(num_samples)


# %%
import os
import Augmentor

dir = 'save_offs/pt2/validation/'

# Get a list of all subdirectories within the main directory
subdirs = [d for d in os.listdir(dir) if os.path.isdir(os.path.join(dir, d))]

# Loop over each subdirectory and apply augmentations
for subdir in subdirs:
    subdir_path = os.path.join(dir, subdir)
    p = Augmentor.Pipeline(subdir_path, output_directory='')
    p.flip_left_right(probability=0.8)
    p.flip_top_bottom(probability=0.8)
    #p.rotate(probability=0.9,max_left_rotation=18,max_right_rotation=18)
    p.rotate90(probability=0.7)
    p.rotate180(probability=0.7)
    p.rotate270(probability=0.7)
    p.scale(probability=0.7, scale_factor=1.1)
    p.random_brightness(probability=0.7,min_factor=0.8,max_factor=1.2)
    p.random_contrast(probability=0.7,min_factor=0.8,max_factor=1.2)

    # Get the number of files in the directory
    num_files = len(os.listdir(subdir_path))

    # Sample additional images if necessary
    num_samples = 200 - num_files
    if num_samples > 0:
        p.sample(num_samples)


# %%


# %%
# Define the paths to your data directories
train_dir = 'save_offs/pt2/train/'
validation_dir = 'save_offs/pt2/validation/'
test_dir = 'save_offs/pt2/test_even/'
test_full_dir = 'save_offs/pt2/test/'


img_width, img_height = 224, 224
BATCHSIZE = 16

# Define a function to preprocess the images
def preprocess_image(image):
    image = tf.image.resize(image, (img_width, img_height))
    image = image / 255.0
    return image


# %%
# Load the train data
train_datagen = keras.preprocessing.image.ImageDataGenerator(
    preprocessing_function=preprocess_image,
    #rotation_range=45,
    #width_shift_range=0.2,
    #height_shift_range=0.2,
    #zoom_range=0.2,
    #horizontal_flip=True,
    #vertical_flip=True
)

val_datagen = keras.preprocessing.image.ImageDataGenerator(
    preprocessing_function=preprocess_image,
    #rotation_range=45,
    #horizontal_flip=True
    )

# %%
train_generator = train_datagen.flow_from_directory(
    train_dir,
    target_size=(img_width, img_height),
    batch_size=BATCHSIZE,
    class_mode='categorical',
    shuffle=True)

# Load the validation data

validation_generator = val_datagen.flow_from_directory(
    validation_dir,
    target_size=(img_width, img_height),
    batch_size=BATCHSIZE,
    class_mode='categorical',
    shuffle=False
)

# Load the test data
test_datagen = keras.preprocessing.image.ImageDataGenerator(preprocessing_function=preprocess_image)
test_generator = test_datagen.flow_from_directory(
    test_dir,
    target_size=(img_width, img_height),
    batch_size=BATCHSIZE,
    class_mode='categorical',
    shuffle=False
)



test_full_datagen = keras.preprocessing.image.ImageDataGenerator(preprocessing_function=preprocess_image)
test_full_generator = test_full_datagen.flow_from_directory(
    test_full_dir,
    target_size=(img_width, img_height),
    batch_size=BATCHSIZE,
    class_mode='categorical',
    shuffle=False
)


# Store the class names in a list
class_names = list(train_generator.class_indices.keys())
n_classes = len(class_names)
print(f'Class names: {class_names}')
print('Num of classes:', n_classes)

# Store the number of images in each set
train_set_size = train_generator.n
validation_set_size = validation_generator.n
test_set_size = test_generator.n
test_full_set_size = test_full_generator.n

print("Train set size:", train_set_size)
print("Validation set size:", validation_set_size)
print("Test set size:", test_set_size)
print('test_full_size:', test_full_set_size)



# %%
model = tf.keras.applications.VGG16(input_shape=(224, 224, 3),
                                   weights = 'imagenet',
                                   include_top = False,
                                   #pooling = 'max'
                                   )

#Add additional layers here

#inputs = tf.keras.Input(shape=(224,224,32))
#X = data_augmentation(inputs)
#Dont dellete this one
X= model.layers[-1].output

# Additonal layers can be added and changed from here
X = tf.keras.layers.GlobalAveragePooling2D()(X)
X = tf.keras.layers.BatchNormalization(axis=-1, momentum=0.99, epsilon=0.001)(X)
X = tf.keras.layers.Dropout(0.1)(X)
#X = tf.keras.layers.Dense(256, activation='relu', kernel_regularizer=regularizers.l1_l2(l1=0.03, l2=0.03), bias_regularizer=regularizers.l1_l2(l1=0.03, l2=0.03))(X)
X = tf.keras.layers.Dense(480, activation='relu', 
                          #kernel_regularizer=regularizers.l2( l2=0.016), 
                          #activity_regularizer=regularizers.l1(l1=0.006), 
                          #bias_regularizer=regularizers.l1(l1=0.006)
                         )(X)
X = tf.keras.layers.BatchNormalization()(X)
X = tf.keras.layers.Dropout(0.6000000000000001)(X)
#X= tf.keras.layers.Flatten()(X)

#softmax activation selects neuron with highest probabity output


predictions = Dense(n_classes, activation="softmax",
                    #kernel_regularizer=regularizers.l2(l2=0.001)
                   )(X)


model_final = Model(model.input, predictions)


# %%
for layers in (model.layers)[:-1]:
    print(layers)
    layers.trainable = False
    
for layer in model_final.layers:
    if isinstance(layer, keras.layers.BatchNormalization):
        layer.trainable = False

# %%
for index, layer in enumerate(model_final.layers):
    print("Layer: {}, Trainable: {}".format(index, layer.trainable))


# %%
model_final.summary()

# %%
checkpoint_path = 'save_offs/butterfly_full_mk1'+date+'.ckpt'
checkpoint_dir = os.path.dirname(checkpoint_path)



# %%
#checkpoint = ModelCheckpoint(monitor='val_accuracy', verbose=1, save_best_only=True, save_weights_only=False, mode='auto', save_freq = 'epochs')
cp_callback = tf.keras.callbacks.ModelCheckpoint(filepath = checkpoint_path,save_best_only = True, monitor='val_accuracy' ,save_weights_only = False, verbose = 1)
early = EarlyStopping(monitor='val_accuracy', min_delta=0.01, patience=2, verbose=4, mode='max')

# %%

base_learning_rate = 0.0001

# %%

# Change loss depending on how many classes there are
        
#base_learning_rate = 0.0001
model_final.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=base_learning_rate),
              loss=tf.keras.losses.categorical_crossentropy,
              metrics=['accuracy',
                       #keras.metrics.SparseTopKCategoricalAccuracy(k=3)
                      ])


# %%
initial_epochs = 5

# %%
def lr_exp_decay(initial_epochs, lr):
    k = 0.01
    return base_learning_rate * math.exp(-k*initial_epochs)

# %%
history = model_final.fit(train_generator,
                          epochs= initial_epochs,
                          validation_data= validation_generator,
                          callbacks=[cp_callback,
                                     early,
                                     #LearningRateScheduler(lr_exp_decay, verbose=1)
                                    ],
                          #batch_size = 128,
                          #shuffle = True
                         )


# %% [markdown]
# 

# %%
test_loss, test_acc = model_final.evaluate(
    test_generator)

# %%
for layer in model_final.layers:
    if isinstance(layer, keras.layers.BatchNormalization):
        layer.trainable = False
    else:
        layer.trainable = True

for index, layer in enumerate(model_final.layers):
    print("Layer: {}, Trainable: {}".format(index, layer.trainable))

# %%
# Let's take a look to see how many layers are in the base model
print("Number of layers in the base model: ", len(model_final.layers))

# Fine-tune from this layer onwards
fine_tune_at = 8

# Freeze all the layers before the `fine_tune_at` layer
for layer in model_final.layers[:fine_tune_at]:
  layer.trainable = False

for layer in model_final.layers:
    if isinstance(layer, keras.layers.BatchNormalization):
        layer.trainable = False


# %%
model_final.summary()

# %%
for index, layer in enumerate(model_final.layers):
    print("Layer: {}, Trainable: {}".format(index, layer.trainable))

# %%
# learning rate has been lowered as more model ahs been opened for training. Should stop overfititng

TL_learningRate = base_learning_rate/10

model_final.compile(optimizer=tf.keras.optimizers.RMSprop(learning_rate=TL_learningRate),
              loss=tf.keras.losses.categorical_crossentropy,
              metrics=['accuracy'])

# %%
fine_tune_epochs = 200

def TF_lr_exp_decay(fine_tune_epochs, lr):
    k = 0.005
    return TL_learningRate * math.exp(-k*fine_tune_epochs)

# %%
total_epochs = initial_epochs + fine_tune_epochs

#np.random.seed(343)
# fit the model
history_fine = model_final.fit(
  train_generator,
  epochs=total_epochs,
  initial_epoch = history.epoch[-1],
  #initial_epoch = fine_tune_epochs,
  validation_data=validation_generator,
  callbacks=[cp_callback, early,
             #LearningRateScheduler(TF_lr_exp_decay, verbose=1)
            ],
  #batch_size=128,
  #shuffle=True
  )


# %%
test_full_loss, test_full_acc = model_final.evaluate(test_full_generator)

print('Test loss:', test_full_loss)
print('test accuracy:', test_full_acc)

# %%
acc_str = str(round(test_acc * 100,2))


model_name = f'save_offs/butt_full_full_acc{acc_str}_{date}.h5'

model_final.save(model_name)

# %%
import pandas as pd

# Make predictions on the test set
predictions = model_final.predict(test_full_generator)

# Convert predictions to class labels
predicted_class_indices = np.argmax(predictions, axis=1)
predicted_classes = [class_names[idx] for idx in predicted_class_indices]

# Get true class labels
true_classes = test_full_generator.classes
class_labels = list(test_full_generator.class_indices.keys())

# Compute confusion matrix and classification report
confusion_mtx = confusion_matrix(true_classes, predicted_class_indices)
class_report = classification_report(true_classes, predicted_class_indices, target_names=class_labels, output_dict=True)

# Convert confusion matrix and classification report to DataFrame
confusion_mtx_df = pd.DataFrame(confusion_mtx, index=class_labels, columns=class_labels)
class_report_df = pd.DataFrame(class_report).transpose()

# Save confusion matrix and classification report to Excel
with pd.ExcelWriter('save_offs/classification_results_test_full.xlsx') as writer:  
    confusion_mtx_df.to_excel(writer, sheet_name='Confusion Matrix')
    class_report_df.to_excel(writer, sheet_name='Classification Report')

# %% [markdown]
# 

# %% [markdown]
# test_loss, test_acc = model_final.evaluate(test_generator)
# 
# print('Test loss:', test_loss)
# print('test accuracy:', test_acc)

# %%
import os
import shutil
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random

# Define directories and other initial setup
misclassified_dir = 'save_offs/misclassified/'
test_dir = '../sorting_dataset/save_offs/pt2/test/'

if not os.path.exists(misclassified_dir):
    os.makedirs(misclassified_dir)

def random_color():
    return "{:02X}{:02X}{:02X}".format(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))

misclassification_details = []
predictions = model_final.predict(test_full_generator)
predicted_classes = np.argmax(predictions, axis=1)
true_classes = test_full_generator.classes
label_to_class_name = {v: k for k, v in test_full_generator.class_indices.items()}
misclassified_idx = np.where(predicted_classes != true_classes)[0]

for idx in misclassified_idx:
    img_filename = test_full_generator.filenames[idx]
    true_label = label_to_class_name[true_classes[idx]]
    predicted_label = label_to_class_name[predicted_classes[idx]]

    misclassification_details.append({
        'Image Filename': os.path.basename(img_filename),
        'True Label': true_label,
        'Predicted Label': predicted_label
    })

    subdir = os.path.dirname(img_filename)
    new_subdir = os.path.join(misclassified_dir, subdir)

    if not os.path.exists(new_subdir):
        os.makedirs(new_subdir)

    src_path = os.path.join(test_dir, img_filename)
    dst_path = os.path.join(new_subdir, os.path.basename(img_filename))

    if os.path.exists(src_path):
        shutil.move(src_path, dst_path)

df = pd.DataFrame(misclassification_details)
excel_path = 'save_offs/misclassification_details.xlsx'
df.to_excel(excel_path, index=False)

book = load_workbook(excel_path)
sheet = book.active

unique_class_names = list(set(df['True Label'].tolist() + df['Predicted Label'].tolist()))
color_mapping = {class_name: random_color() for class_name in unique_class_names}

for row in range(2, sheet.max_row + 1):
    true_label_cell = f'B{row}'
    predicted_label_cell = f'C{row}'
    true_label = sheet[true_label_cell].value
    predicted_label = sheet[predicted_label_cell].value
    true_label_color = color_mapping.get(true_label, 'FFFFFF')
    predicted_label_color = color_mapping.get(predicted_label, 'FFFFFF')
    fill_true = PatternFill(start_color=true_label_color, end_color=true_label_color, fill_type='solid')
    fill_predicted = PatternFill(start_color=predicted_label_color, end_color=predicted_label_color, fill_type='solid')
    sheet[true_label_cell].fill = fill_true
    sheet[predicted_label_cell].fill = fill_predicted

book.save(excel_path)


# %%
# Define source directory and target directory
src_dir = "save_offs/"
base_target_dir = "../../../../../media/cryptobiovis/17d4ec92-0c5a-4c67-8495-32dbeae14c24/sorted_17_oct_save_offs"

# Add the iteration number to the folder name
target_dir = f"{base_target_dir}_{iteration}"

# Check if target directory exists
if os.path.exists(target_dir):
    print(f"Target directory {target_dir} already exists. Please check.")
else:
    # Move the directory
    shutil.move(src_dir, target_dir)

# %%


# %%


# %%



